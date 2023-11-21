import openpyxl
from openpyxl.styles import Alignment, PatternFill

def main():
    # load the excel file
    file_path = 'Book2.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    # select a specific worksheet
    worksheet = workbook['Sheet1']

    # create an array of cell columns used
    cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
            'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
            'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ',
            'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT',
            'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD',
            'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL']
    
    days = ['20', '21', '22', '23', '24', '25', '26', '27', '28', '29',
             '30', '31', '1', '2', '3', '4', '5', '6', '7', '8', '9',
             '10', '11', '12', '13', '14', '15', '16', '17', '18', '19',
            ]

    # modify cell values
    worksheet.cell(row=4, column=2, value="Nome")

    # how to join two cells
    # Assuming you want to join cells A1 and B1 and store the result in A1
    worksheet.merge_cells('A1:B1')

    # loop through the array and merge the values
    count = 0
    day_num = 0
    col_num = 3
    num = 2
    for col in range(len(cols)):
        worksheet.merge_cells(f"{cols[count]}{str(num)}:{cols[count+1]}{str(num)}")
        merged_cell = worksheet[f"{cols[count]}{str(num)}"]
        cell = worksheet[f"{cols[count]}{str(num)}"]
        alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.alignment = alignment
        fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.fill = fill_color


        if count < 60:
            count += 2
        
        col_num += 2

    # add the days
    col_num = 3    
    for day in days:
        worksheet.cell(row=2, column=col_num, value=days[day_num])
        day_num += 1
        col_num += 2

    # add the texts entrada e saida respectively
    en = 0
    en_col = 3
    for col in range(len(days)):
        worksheet.cell(row=3, column=en_col, value="Entrada")
        worksheet.cell(row=3, column=en_col + 1, value="Saída")

        en_col += 2

    # create lists contaning each person from each company
    angobv = ["Ilidio Quimuanga 1", "Manuel Henriques 1", "Nguinamau Pedro 1", "Hermenegildo Jose 1", "Mubua Beloti 1", "Bernardo Adao 1", "Antonio Miranda 1",
              "Fernando de Almeid 1", "Erico dos Santos 1", "Neusa Sardinha 1", "Rui de Almeida 1", "Cesaltina de Sousa 1", "Mario Jardel 1", "Ernesto Isabel 1",
              "Madalena Andre 1", "Ventura Alberto 1", "Conceicao Malungo 1", "Nlandu Amulzila 1", "Maria Simao 1", "Alberta Bernardo 1", "Chelseo Leonel 1"] 

    angobavaria = ["ilidio Mendes Quimuanga", "Manuel Miguel Henriques", "Nguinamau Pedro Tiago",
                   "Hermenegildo Jose Fonsceca", "Mubua Quinvuita Beloti", "Bernardo Januario Adao",
                   "Antonio Francisco Miranda Mussique", "Fernando de Almeida", "Erico Diocleciano Manuel dos Santos",
                   "Neusa Julio Sardinha", "Rui Eduardo Mariano de Almeida", "Cesaltina Caunji de Sousa Bonifacio",
                   "Mario Jardel MIguel dos Santos", "Ernesto Isabel", "Madalena Teresa Andre", "Ventura Alberto Samuzaria Saicosse",
                   "Conceicao Miranda Ribeiro Malungo", "Nlandu Amunzila", "Maria da Conceição Simão", "Alberta Bernardo", "Chelseo Leonel Miguel Sebastiao"]
    

    ab_num = 0
    ab_row = 4
    for ab in angobavaria:
        worksheet.cell(row=ab_row, column=2, value=angobavaria[ab_num])
        ab_num += 1
        ab_row += 1
    
    # check to see if you can read from the excel file
    ws = workbook['Sheet2']
    d = ws['c1']
    print(d.value)
    print(ws['c1'].value)

    # read all the names and add them to a list
    # then remove the duplicates
    cell_row = 0
    cell_column = 0
    names = []

    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        # 'min_row=2' skips the header row, assuming it's in the first row
        # 'values_only=True' ensures you get the cell values instead of Cell objects
        for cell_value in row:
            # Do something with the cell_value
            names.append(cell_value)
            # print(cell_value)
    new_names = list(dict.fromkeys(names))
    print(new_names)

    cell_count = 0
    row_count = 1
    col_count = 0
    timeStamps = []
    for name in names:
        if new_names[cell_count] == ws[f"A{row_count}"].value:
            print("true")
            timeStamps.append(str(ws[f"C{row_count}"].value))
        row_count += 1


    print(timeStamps)
    print(len(timeStamps))




    # save the changes
    workbook.save('Book2.xlsx')

    # close the workbook
    workbook.close()

    # code to remove duplicate values from a list
    # my_list = ["mobile","laptop","earphones","mobile", 'laptop']
    # new_list = list(set(my_list))
    # print(f"the old list was {my_list}, the new list without duplicates is {new_list}")
    print("success")
    

if __name__ == "__main__":
    main()
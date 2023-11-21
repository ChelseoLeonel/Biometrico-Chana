import openpyxl

def main():
    # load the excel file
    file_path = 'test.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    # select a specific worksheet
    worksheet = workbook['Sheet1']

    # modify cell values
    worksheet.cell(row=1, column=1, value="Chelseo Miguel")

    # how to join two cells
    worksheet.merge_cells('A2:D2')
    worksheet.unmerge_cells('A2:D2')
    # or equivalently
    worksheet.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
    worksheet.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)


    # save the changes
    workbook.save('test.xlsx')

    # close the workbook
    workbook.close()

if __name__ == "__main__":
    main()


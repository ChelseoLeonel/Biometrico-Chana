import openpyxl
from openpyxl.styles import Alignment, PatternFill

def main():
    # load the excel file
    file_path = 'Book2.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    # select a specific worksheet
    worksheet = workbook['Sheet1']

    # create an array of cell columns used
    cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
            'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
            'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ',
            'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT',
            'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD',
            'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL']
    
    days = ['20', '21', '22', '23', '24', '25', '26', '27', '28', '29',
             '30', '31', '01', '02', '03', '04', '05', '06', '07', '08', '09',
             '10', '11', '12', '13', '14', '15', '16', '17', '18', '19',
            ]

    # modify cell values
    worksheet.cell(row=3, column=2, value="Nome")

    # how to join two cells
    # Assuming you want to join cells A1 and B1 and store the result in A1
    worksheet.merge_cells('A1:B1')

    # loop through the array and merge the values
    count = 0
    day_num = 0
    col_num = 3
    num = 2
    for col in range(len(cols)):
        worksheet.merge_cells(f"{cols[count]}{str(num)}:{cols[count+1]}{str(num)}")
        merged_cell = worksheet[f"{cols[count]}{str(num)}"]
        cell = worksheet[f"{cols[count]}{str(num)}"]
        alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.alignment = alignment
        fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.fill = fill_color


        if count < 60:
            count += 2
        
        col_num += 2

    # add the days
    col_num = 3    
    for day in days:
        worksheet.cell(row=2, column=col_num, value=days[day_num])
        day_num += 1
        col_num += 2

    # add the texts entrada e saida respectively
    en = 0
    en_col = 3
    for col in range(len(days)):
        worksheet.cell(row=3, column=en_col, value="Entrada")
        worksheet.cell(row=3, column=en_col + 1, value="Saída")

        en_col += 2

    # create lists contaning each person from each company
    angobv = ["Ilidio Quimuanga 1", "Manuel Henriques 1", "Nguinamau Pedro 1", "Hermenegildo Jose 1", "Mubua Beloti 1", "Bernardo Adao 1", "Antonio Miranda 1",
              "Fernando de Almeid 1", "Erico dos Santos 1", "Neusa Sardinha 1", "Rui de Almeida 1", "Cesaltina de Sousa 1", "Mario Jardel 1", "Ernesto Isabel 1",
              "Madalena Andre 1", "Ventura Alberto 1", "Conceicao Malungo 1", "Nlandu Amulzila 1", "Maria Simao 1", "Alberta Bernardo 1", "Chelseo Leonel 1"] 

    angobavaria = ["ilidio Mendes Quimuanga", "Manuel Miguel Henriques", "Nguinamau Pedro Tiago",
                   "Hermenegildo Jose Fonsceca", "Mubua Quinvuita Beloti", "Bernardo Januario Adao",
                   "Antonio Francisco Miranda Mussique", "Fernando de Almeida", "Erico Diocleciano Manuel dos Santos",
                   "Neusa Julio Sardinha", "Rui Eduardo Mariano de Almeida", "Cesaltina Caunji de Sousa Bonifacio",
                   "Mario Jardel MIguel dos Santos", "Ernesto Isabel", "Madalena Teresa Andre", "Ventura Alberto Samuzaria Saicosse",
                   "Conceicao Miranda Ribeiro Malungo", "Nlandu Amunzila", "Maria da Conceição Simão", "Alberta Bernardo", "Chelseo Leonel Miguel Sebastiao"]
    
    rac = ["Ciclena da Silva 1", "Eduarda Alves 1", "Emanuel Soares 1", "Jeronimo Alfredo 1", "Jose Dombo 1", "Josemar Francisco 1", "Lucia da Cruz 1", "Manuel Pacheco 1",
           "Miguens Antonio 1", "Paulino Ngunji 1", "Regina Ernesto 1", "Isoveth Alfredo 1", "Augusto Mota 1", "Adilson Camate 1", "Ana Dulo 1"]

    rentacar = ["Ciclena Fernando da Silva", "Eduarda Bartolomeu Alves", "Emanuel buchartts Ramos Soares",
                "Jeronimo Domingos Alfredo", "Jose Americo Dombo", "Josemar Manzambi Celestino Francisco", "Lucia de Jesus Certo Kumbi da Cruz",
                "Manuel Pacheco Luis Cardoso", "Miguens Antonio Marcolino", "Paulino Ngunji", "Regina Ferreira Ernesto",
                "Isoveth Alfredo Issunge", "Augusto Helder Vumbi Mota", "Adilson Antonio Satota Camate", "Ana Zanatt Coelho Dulo"]
    
    cSA = ["Cristina Quingombo 1", "Pedro Manuel 1", "Ezildo Joao 1", "Eliane Cortez 1", "Marcia Banda 1", "Isabel Virgilio 1",
           "Lisandra Fontoura 1", "Celio Marcos 1"]

    chanaSA = ["Cristina Sebastiao Quingombo Candeeiro", "Pedro Jose Manuel", "Ezildo da Conceicao Joao", "Eliane Ferreira Balde Alfa Cortez", "Marcia da Consolocao Constantino dos Santos",
               "Isabel Virgilio", "Lisandra Sofia Henriques Simao Fontoura", "Celio Adilson Pedro Marcos"]
    
    ap = ["Antonica Alfredo 1", "Inacio Zaila 1", "Joana Chipenda 1", "Jose Capassera 1", "Julia Guia 1", "Madalena Antonio 1", "Eduardo Dinis 1", "Ana Manuel 1"]

    auto_p = ["Antonica Jose Alfredo", "Inacio Kansobo Clemente Zaila", "Ermelinda Joana Kopumi Chipenda", "Jose Candeeiro Capassera", "Julia de Fatima Manuel Guia", "Madalena Malungo Antonio",
              "Eduardo Dinis Manuel", "Ana Teresa Braganca Manuel"]
    
    soss = ["Danilson  Andre 1", "Joaquim Freitas 1", "Sanzu Bernardo 1", "Malungo Rafael 1", "Daniel  Joao 1", "Kamalando Antonio 1", "Afonso Antonio 1", "Henriques Nganga 1",
            "Romao  Zenduca 1", "Dinora Santos 1", "Alfredo Noy 1", "Fernandes Julio 1", "Edna de Sousa 1", "Agostinho Vieira 1", "Lemos Victor 1", "Moises Gumba 1",
            "Domingas Vungi 1", "Antonio Barros 1", "Muriel Junqueira", "Eliseu Joao 1", "Germano Gonga 1", "Cilio Miguel 1", "Antonio Pires 1", "Kiteque Manuel 1", "Domingas Fragoso 1", "Armando Augusto 1", "Pedro Guerra 1"]
    
    sossego = ["Danilson Jaime Torres Andre", "Joaquim Jorge Correia Freitas", "Sanzu Antonio Paulo Bernardo", "Malungo Rafael", "Daniel Francisco Joao", "Kamalando Antonio", "Afonso Pedro Antonio", "Henriques Luis Nganga",
                "Romao Mauricio Zenduca", "Dinora da Graça Bernado dos Santos", "Alfredo Armando Noy", "Fernandes Filipe Júlio", "Edna Teresa Simoes Cristovao de Sousa", "Agostinho Antonio Vieira",
                "Lemos Alberto Victor", "Moises Antonio Gumba","Domingas Angela Sebastiao Vungi Joao", "Antonio Jose Amaro de Barros", "Muriel da Conceicao Junqueira", "Eliseu Francisco Joao",
                "Germano Daniel Gonga", "Cilio Dirinelo Miguel Sebastiao", "Antonio Simoes Pires", "Kiteque Domingos Kitumba Manuel", "Domingas Cireucia Vapor Fragoso", "Armando Luis Francony Augusto", "Pedro Guerra"]
    

    ab_num = 0
    ab_row = 4
    for ab in angobv:
        worksheet.cell(row=ab_row, column=2, value=angobavaria[ab_num])
        ab_num += 1
        ab_row += 1
    
    # check to see if you can read from the excel file
    ws = workbook['Sheet2']
    d = ws['c1']
    print(d.value)
    print(ws['c1'].value)

    # read all the names and add them to a list
    # then remove the duplicates
    cell_row = 0
    cell_column = 0
    names = []

    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        # 'min_row=2' skips the header row, assuming it's in the first row
        # 'values_only=True' ensures you get the cell values instead of Cell objects
        for cell_value in row:
            # Do something with the cell_value
            names.append(cell_value)
            # print(cell_value)
    new_names = list(dict.fromkeys(names))
    print(new_names)

    cell_count = 0
    row_count = 1
    col_count = 0
    timeStamps = []

    # for angobavaria
    p_row = 4
    p_col = 3
    p_col2 = 4
    p_index = 0

    letters = ["C2", "E2", "G2", "I2", "K2", "M2", "O2", "Q2", "S2", "U2", "W2", "Y2", "AA2", "AC2",
               "AE2", "AG2", "AI2", "AK2", "AM2", "AO2", "AQ2", "AS2", "AU2", "AW2", "AY2", "BA2",
               "BC2", "BE2", "BG2", "BI2", "BK2"]
    letters_count = 0
    v_row = 0
    v_col = 0
    for person in angobv:
        for name in names:
            if person == ws[f"A{row_count}"].value:
                print("true")
                timeStamps.append(str(ws[f"C{row_count}"].value))
            row_count += 1
            # print time stamp to check the format 
            #print(timeStamps)
        for time in timeStamps:
            for letter in letters:
                if time[8:10] == worksheet[letter].value:
                    if int(time[11:13]) < 11:
                        worksheet.cell(row=p_row, column=p_col, value=time)
                    else:
                        worksheet.cell(row=p_row, column=p_col + 1, value=time)
                else:
                    p_col += 2
            p_col = 3
        p_col = 3
        p_row += 1
        row_count = 1
        timeStamps = []


    print(timeStamps)
    print(len(timeStamps))




    # save the changes
    workbook.save('Book2.xlsx')

    # close the workbook
    workbook.close()

    # code to remove duplicate values from a list
    # my_list = ["mobile","laptop","earphones","mobile", 'laptop']
    # new_list = list(set(my_list))
    # print(f"the old list was {my_list}, the new list without duplicates is {new_list}")
    print("success")
    

if __name__ == "__main__":
    main()